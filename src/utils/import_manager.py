"""
Import utilities - CSV/Excel bestanden inlezen
"""

import csv
import os
from src.models.klant import Klant

class ImportManager:
    """Beheert import van klanten uit CSV/Excel"""
    
    # Kolom mapping van jouw Excel naar ons model
    COLUMN_MAPPING = {
        'Nummer': 'klantnummer',
        'Naam': 'naam',
        'Telefoon': 'telefoon',
        'E-mailadres': 'email',
        'Adres': 'straat',
        'Postcode': 'postcode',
        'Plaats': 'gemeente',
        'GSM nummer': 'gsm',
        'contactpersoon': 'contactpersoon',
    }
    
    @staticmethod
    def import_csv(file_path):
        """
        Importeert klanten uit CSV bestand
        
        Args:
            file_path: Pad naar CSV bestand
            
        Returns:
            tuple: (lijst met Klant objecten, lijst met fouten)
        """
        klanten = []
        fouten = []
        
        if not os.path.exists(file_path):
            fouten.append(f"Bestand niet gevonden: {file_path}")
            return klanten, fouten
        
        try:
            # Detecteer delimiter (tab of comma)
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                first_line = f.readline()
                delimiter = '\t' if '\t' in first_line else ','
            
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f, delimiter=delimiter)
                
                if reader.fieldnames is None:
                    fouten.append("CSV bestand is leeg of ongeldig")
                    return klanten, fouten
                
                for row_num, row in enumerate(reader, start=2):  # Start bij 2 (header is 1)
                    try:
                        # Maak Klant object met relevante velden
                        klant = Klant()
                        
                        # Map velden van Excel naar Klant
                        for excel_col, klant_attr in ImportManager.COLUMN_MAPPING.items():
                            if excel_col in row:
                                value = row[excel_col].strip() if row[excel_col] else None
                                if value:
                                    setattr(klant, klant_attr, value)
                        
                        # Validatie: klantnummer en naam zijn verplicht
                        if not klant.klantnummer:
                            fouten.append(f"Rij {row_num}: Klantnummer ontbreekt")
                            continue
                        
                        if not klant.naam:
                            fouten.append(f"Rij {row_num}: Naam ontbreekt")
                            continue
                        
                        klanten.append(klant)
                        
                    except Exception as e:
                        fouten.append(f"Rij {row_num}: {str(e)}")
        
        except Exception as e:
            fouten.append(f"Fout bij lezen bestand: {str(e)}")
        
        return klanten, fouten
    
    @staticmethod
    def import_excel(file_path):
        """
        Importeert klanten uit Excel bestand (.xlsx)
        
        Args:
            file_path: Pad naar Excel bestand
            
        Returns:
            tuple: (lijst met Klant objecten, lijst met fouten)
        """
        try:
            import openpyxl
        except ImportError:
            return [], ["openpyxl niet geïnstalleerd. Installeer met: pip install openpyxl"]
        
        klanten = []
        fouten = []
        
        if not os.path.exists(file_path):
            fouten.append(f"Bestand niet gevonden: {file_path}")
            return klanten, fouten
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Lees headers (eerste rij)
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            if not headers or headers[0] is None:
                fouten.append("Excel bestand is leeg of ongeldig")
                return klanten, fouten
            
            # Lees data rijen
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Maak dictionary van header -> value
                    row_data = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                    
                    # Maak Klant object
                    klant = Klant()
                    
                    # Map velden
                    for excel_col, klant_attr in ImportManager.COLUMN_MAPPING.items():
                        if excel_col in row_data:
                            value = row_data[excel_col]
                            if isinstance(value, str):
                                value = value.strip()
                            if value:
                                setattr(klant, klant_attr, str(value))
                    
                    # Validatie
                    if not klant.klantnummer:
                        fouten.append(f"Rij {row_num}: Klantnummer ontbreekt")
                        continue
                    
                    if not klant.naam:
                        fouten.append(f"Rij {row_num}: Naam ontbreekt")
                        continue
                    
                    klanten.append(klant)
                    
                except Exception as e:
                    fouten.append(f"Rij {row_num}: {str(e)}")
            
            wb.close()
            
        except Exception as e:
            fouten.append(f"Fout bij lezen Excel: {str(e)}")
        
        return klanten, fouten
